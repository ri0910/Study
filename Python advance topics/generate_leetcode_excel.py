import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =============================
# 1️⃣ Fetch problem data from LeetCode API
# =============================
print("📡 Fetching problem data from LeetCode...")

url = "https://leetcode.com/api/problems/all/"
response = requests.get(url)
data = response.json()

problems_data = data["stat_status_pairs"]

problems = []
for item in problems_data:
    stat = item["stat"]
    difficulty = item["difficulty"]["level"]
    paid_only = item["paid_only"]
    if paid_only:
        continue  # skip paid problems
    title = stat["question__title"]
    title_slug = stat["question__title_slug"]

    if difficulty == 1:
        diff = "Easy"
    elif difficulty == 2:
        diff = "Medium"
    else:
        diff = "Hard"

    link = f"https://leetcode.com/problems/{title_slug}/"
    problems.append(["General", title, diff, link])

print(f"✅ Total problems fetched: {len(problems)}")

# =============================
# 2️⃣ Convert to DataFrame and limit to top 200
# =============================
df = pd.DataFrame(problems, columns=["Topic", "Problem Title", "Difficulty", "Link"])
df = df.sort_values(by="Difficulty", key=lambda x: x.map({"Easy": 1, "Medium": 2, "Hard": 3}))
df = df.head(200)
df["Status"] = "☐"

# =============================
# 3️⃣ Save to Excel
# =============================
excel_path = "Top_200_LeetCode_DSA_Problems.xlsx"
df.to_excel(excel_path, index=False, sheet_name="DSA_Problems")

# =============================
# 4️⃣ Beautify Excel sheet
# =============================
wb = load_workbook(excel_path)
ws = wb.active

header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
header_font = Font(bold=True)
for col_num, col_name in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    ws.column_dimensions[get_column_letter(col_num)].width = 30

# Align text
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center")

wb.save(excel_path)
print(f"🎯 Excel file created successfully: {excel_path}")
